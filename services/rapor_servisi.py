import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Turkce karakter destegi icin DejaVu Sans fontu
pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
PDF_FONT = 'DejaVuSans'
PDF_FONT_BOLD = 'DejaVuSans-Bold'

from models import Daire, Odeme, Gider, GiderKalemi, Kasa, Ayar, AidatAyari

AY_ISIMLERI = {
    1: 'Ocak', 2: 'Subat', 3: 'Mart', 4: 'Nisan', 5: 'Mayis', 6: 'Haziran',
    7: 'Temmuz', 8: 'Agustos', 9: 'Eylul', 10: 'Ekim', 11: 'Kasim', 12: 'Aralik'
}

# ── Excel Styles ──────────────────────────────────────────────
BASLIK_FONT = Font(bold=True, size=14, color='FFFFFF')
ALT_BASLIK_FONT = Font(bold=True, size=11, color='FFFFFF')
BASLIK_FILL = PatternFill(start_color='1a1d21', end_color='1a1d21', fill_type='solid')
YESIL_FILL = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
KIRMIZI_FILL = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
INCE_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# ── Excel Helpers ─────────────────────────────────────────────
def _excel_baslik(ws, apartman_adi, rapor_adi, tarih_str):
    ws.append([apartman_adi])
    ws.append([rapor_adi])
    ws.append([f'Olusturma Tarihi: {tarih_str}'])
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=1):
        for cell in row:
            cell.font = BASLIK_FONT if cell.row == 1 else ALT_BASLIK_FONT
            cell.fill = BASLIK_FILL
            cell.alignment = Alignment(horizontal='left')


def _excel_tablo_basligi(ws, basliklar, satir_no):
    for col_idx, baslik in enumerate(basliklar, 1):
        cell = ws.cell(row=satir_no, column=col_idx, value=baslik)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = BASLIK_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = INCE_BORDER


def _excel_satir(ws, degerler, satir_no, fill=None):
    for col_idx, deger in enumerate(degerler, 1):
        cell = ws.cell(row=satir_no, column=col_idx, value=deger)
        cell.border = INCE_BORDER
        cell.alignment = Alignment(horizontal='center')
        if fill:
            cell.fill = fill


def _kaydet_excel(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── PDF Helpers ───────────────────────────────────────────────
def _pdf_baslik_tablosu(apartman_adi, rapor_adi, tarih_str):
    title_style = ParagraphStyle('TRTitle', fontName=PDF_FONT_BOLD, fontSize=16, spaceAfter=4)
    heading_style = ParagraphStyle('TRHeading', fontName=PDF_FONT_BOLD, fontSize=12, spaceAfter=4)
    normal_style = ParagraphStyle('TRNormal', fontName=PDF_FONT, fontSize=9, textColor=colors.grey)
    elements = []
    elements.append(Paragraph(apartman_adi, title_style))
    elements.append(Paragraph(rapor_adi, heading_style))
    elements.append(Paragraph(f'Olusturma Tarihi: {tarih_str}', normal_style))
    elements.append(Spacer(1, 10 * mm))
    return elements


def _pdf_tablo(basliklar, veriler, col_widths=None):
    data = [basliklar] + veriler
    t = Table(data, colWidths=col_widths)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    # Alternating row colors
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f0f0f0')))
    t.setStyle(TableStyle(style_cmds))
    return t


def _kaydet_pdf(elements):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    doc.build(elements)
    output.seek(0)
    return output


def _apartman_adi():
    return Ayar.getir('apartman_adi', 'Apartman')


def _tarih_str():
    return datetime.now().strftime('%d.%m.%Y')


def _ay_hesapla(yil, ay):
    """Bir ay icin gelir/gider/gider dagilimini dogrudan hesapla (kasaya bagimli degil)."""
    aidat = AidatAyari.guncel_miktar()
    odenen = Odeme.query.filter_by(yil=yil, ay=ay, odendi=True).count()
    toplam_daire = Daire.query.count()
    gelir = odenen * aidat

    giderler = Gider.query.filter_by(yil=yil, ay=ay).all()
    gider_toplam = sum(g.tutar for g in giderler)

    kalem_toplam = {}
    for g in giderler:
        ad = g.kalem.kalem_adi
        kalem_toplam[ad] = kalem_toplam.get(ad, 0) + g.tutar

    return {
        'aidat': aidat,
        'odenen': odenen,
        'toplam_daire': toplam_daire,
        'gelir': gelir,
        'gider': gider_toplam,
        'kalem_toplam': kalem_toplam,
    }


# ══════════════════════════════════════════════════════════════
# 1) Aylik Ozet
# ══════════════════════════════════════════════════════════════
def aylik_ozet_excel(yil, ay):
    apartman = _apartman_adi()
    ay_adi = AY_ISIMLERI.get(ay, '')
    tarih = _tarih_str()
    h = _ay_hesapla(yil, ay)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Aylik Ozet'

    _excel_baslik(ws, apartman, f'{ay_adi} {yil} - Aylik Ozet Raporu', tarih)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    satir = 5
    bilgiler = [
        ('Aidat Tutari', f'{h["aidat"]:.2f} TL'),
        ('Odeme Yapan', f'{h["odenen"]} / {h["toplam_daire"]}'),
        ('Toplam Gelir', f'{h["gelir"]:.2f} TL'),
        ('Toplam Gider', f'{h["gider"]:.2f} TL'),
        ('Net', f'{h["gelir"] - h["gider"]:.2f} TL'),
    ]
    for etiket, deger in bilgiler:
        _excel_satir(ws, [etiket, deger], satir)
        satir += 1

    if h['kalem_toplam']:
        satir += 1
        _excel_tablo_basligi(ws, ['Gider Kalemi', 'Tutar'], satir)
        satir += 1
        for ad, tutar in h['kalem_toplam'].items():
            _excel_satir(ws, [ad, f'{tutar:.2f} TL'], satir)
            satir += 1
        _excel_satir(ws, ['TOPLAM', f'{h["gider"]:.2f} TL'], satir)
        ws.cell(row=satir, column=1).font = Font(bold=True)
        ws.cell(row=satir, column=2).font = Font(bold=True)

    return _kaydet_excel(wb)


def aylik_ozet_pdf(yil, ay):
    apartman = _apartman_adi()
    ay_adi = AY_ISIMLERI.get(ay, '')
    tarih = _tarih_str()
    h = _ay_hesapla(yil, ay)

    elements = _pdf_baslik_tablosu(apartman, f'{ay_adi} {yil} - Aylik Ozet Raporu', tarih)

    ozet_data = [
        ['Aidat Tutari', f'{h["aidat"]:.2f} TL'],
        ['Odeme Yapan', f'{h["odenen"]} / {h["toplam_daire"]}'],
        ['Toplam Gelir', f'{h["gelir"]:.2f} TL'],
        ['Toplam Gider', f'{h["gider"]:.2f} TL'],
        ['Net', f'{h["gelir"] - h["gider"]:.2f} TL'],
    ]
    elements.append(_pdf_tablo(['Bilgi', 'Deger'], ozet_data, col_widths=[60 * mm, 60 * mm]))
    elements.append(Spacer(1, 10 * mm))

    if h['kalem_toplam']:
        gider_rows = [[ad, f'{tutar:.2f} TL'] for ad, tutar in h['kalem_toplam'].items()]
        gider_rows.append(['TOPLAM', f'{h["gider"]:.2f} TL'])
        elements.append(_pdf_tablo(['Gider Kalemi', 'Tutar'], gider_rows, col_widths=[60 * mm, 60 * mm]))

    return _kaydet_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 2) Odeme Durumu
# ══════════════════════════════════════════════════════════════
def odeme_durumu_excel(yil, ay):
    apartman = _apartman_adi()
    ay_adi = AY_ISIMLERI.get(ay, '')
    tarih = _tarih_str()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Odeme Durumu'

    _excel_baslik(ws, apartman, f'{ay_adi} {yil} - Odeme Durumu Raporu', tarih)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    satir = 5
    _excel_tablo_basligi(ws, ['Daire No', 'Sakin Adi', 'Durum'], satir)
    satir += 1

    daireler = Daire.query.order_by(Daire.daire_no).all()
    odenen_sayisi = 0
    for d in daireler:
        odeme = Odeme.query.filter_by(daire_id=d.id, yil=yil, ay=ay).first()
        odendi = odeme.odendi if odeme else False
        durum = 'Odendi' if odendi else 'Odenmedi'
        fill = YESIL_FILL if odendi else KIRMIZI_FILL
        _excel_satir(ws, [d.daire_no, d.sakin_adi, durum], satir, fill=fill)
        if odendi:
            odenen_sayisi += 1
        satir += 1

    satir += 1
    ws.cell(row=satir, column=1, value='Toplam Odenen:').font = Font(bold=True)
    ws.cell(row=satir, column=2, value=f'{odenen_sayisi} / {len(daireler)}')

    return _kaydet_excel(wb)


def odeme_durumu_pdf(yil, ay):
    apartman = _apartman_adi()
    ay_adi = AY_ISIMLERI.get(ay, '')
    tarih = _tarih_str()

    elements = _pdf_baslik_tablosu(apartman, f'{ay_adi} {yil} - Odeme Durumu Raporu', tarih)

    daireler = Daire.query.order_by(Daire.daire_no).all()
    rows = []
    odenen_sayisi = 0
    for d in daireler:
        odeme = Odeme.query.filter_by(daire_id=d.id, yil=yil, ay=ay).first()
        odendi = odeme.odendi if odeme else False
        durum = 'Odendi' if odendi else 'Odenmedi'
        rows.append([str(d.daire_no), durum])
        if odendi:
            odenen_sayisi += 1

    basliklar = ['Daire No', 'Durum']
    data = [basliklar] + rows
    col_widths = [40 * mm, 50 * mm]
    t = Table(data, colWidths=col_widths)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    for i in range(1, len(data)):
        row_data = rows[i - 1]
        if row_data[1] == 'Odendi':
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#d4edda')))
        else:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8d7da')))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    elements.append(Spacer(1, 5 * mm))
    ozet_style = ParagraphStyle('TROzet', fontName=PDF_FONT_BOLD, fontSize=10)
    elements.append(Paragraph(f'Toplam Odenen: {odenen_sayisi} / {len(daireler)}', ozet_style))

    return _kaydet_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 3) Gider Detay
# ══════════════════════════════════════════════════════════════
def gider_detay_excel(yil, ay):
    apartman = _apartman_adi()
    ay_adi = AY_ISIMLERI.get(ay, '')
    tarih = _tarih_str()
    h = _ay_hesapla(yil, ay)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gider Detay'

    _excel_baslik(ws, apartman, f'{ay_adi} {yil} - Gider Detay Raporu', tarih)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12

    satir = 5
    _excel_tablo_basligi(ws, ['Gider Kalemi', 'Tutar', 'Oran (%)'], satir)
    satir += 1
    for ad, tutar in h['kalem_toplam'].items():
        oran = (tutar / h['gider'] * 100) if h['gider'] > 0 else 0
        _excel_satir(ws, [ad, f'{tutar:.2f} TL', f'{oran:.1f}%'], satir)
        satir += 1
    _excel_satir(ws, ['TOPLAM', f'{h["gider"]:.2f} TL', '100%'], satir)
    for col in range(1, 4):
        ws.cell(row=satir, column=col).font = Font(bold=True)

    return _kaydet_excel(wb)


def gider_detay_pdf(yil, ay):
    apartman = _apartman_adi()
    ay_adi = AY_ISIMLERI.get(ay, '')
    tarih = _tarih_str()
    h = _ay_hesapla(yil, ay)

    elements = _pdf_baslik_tablosu(apartman, f'{ay_adi} {yil} - Gider Detay Raporu', tarih)

    rows = []
    for ad, tutar in h['kalem_toplam'].items():
        oran = (tutar / h['gider'] * 100) if h['gider'] > 0 else 0
        rows.append([ad, f'{tutar:.2f} TL', f'{oran:.1f}%'])
    rows.append(['TOPLAM', f'{h["gider"]:.2f} TL', '100%'])

    elements.append(_pdf_tablo(['Gider Kalemi', 'Tutar', 'Oran (%)'], rows,
                               col_widths=[50 * mm, 40 * mm, 30 * mm]))

    return _kaydet_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 4) Yillik Ozet
# ══════════════════════════════════════════════════════════════
def yillik_ozet_excel(yil):
    apartman = _apartman_adi()
    tarih = _tarih_str()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Yillik Ozet'

    _excel_baslik(ws, apartman, f'{yil} - Yillik Ozet Raporu', tarih)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    satir = 5
    _excel_tablo_basligi(ws, ['Ay', 'Gelir', 'Gider', 'Net'], satir)
    satir += 1

    toplam_gelir = 0
    toplam_gider = 0
    for ay_no in range(1, 13):
        h = _ay_hesapla(yil, ay_no)
        toplam_gelir += h['gelir']
        toplam_gider += h['gider']
        net = h['gelir'] - h['gider']
        ay_adi = AY_ISIMLERI[ay_no]
        _excel_satir(ws, [ay_adi, f'{h["gelir"]:.2f} TL', f'{h["gider"]:.2f} TL',
                          f'{net:.2f} TL'], satir)
        satir += 1
    net_toplam = toplam_gelir - toplam_gider
    _excel_satir(ws, ['TOPLAM', f'{toplam_gelir:.2f} TL', f'{toplam_gider:.2f} TL',
                      f'{net_toplam:.2f} TL'], satir)
    for col in range(1, 5):
        ws.cell(row=satir, column=col).font = Font(bold=True)

    return _kaydet_excel(wb)


def yillik_ozet_pdf(yil):
    apartman = _apartman_adi()
    tarih = _tarih_str()

    elements = _pdf_baslik_tablosu(apartman, f'{yil} - Yillik Ozet Raporu', tarih)

    rows = []
    toplam_gelir = 0
    toplam_gider = 0
    for ay_no in range(1, 13):
        h = _ay_hesapla(yil, ay_no)
        toplam_gelir += h['gelir']
        toplam_gider += h['gider']
        net = h['gelir'] - h['gider']
        ay_adi = AY_ISIMLERI[ay_no]
        rows.append([ay_adi, f'{h["gelir"]:.2f} TL', f'{h["gider"]:.2f} TL',
                     f'{net:.2f} TL'])
    net_toplam = toplam_gelir - toplam_gider
    rows.append(['TOPLAM', f'{toplam_gelir:.2f} TL', f'{toplam_gider:.2f} TL',
                 f'{net_toplam:.2f} TL'])

    elements.append(_pdf_tablo(['Ay', 'Gelir', 'Gider', 'Net'], rows,
                               col_widths=[35 * mm, 40 * mm, 40 * mm, 40 * mm]))

    return _kaydet_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 5) Daire Bazli Odeme Raporu
# ══════════════════════════════════════════════════════════════
def daire_rapor_pdf(daire_id, yil):
    apartman = _apartman_adi()
    tarih = _tarih_str()
    daire = Daire.query.get(daire_id)
    aidat = AidatAyari.guncel_miktar()

    kat_str = 'Giris Kat' if daire.kat == 0 else f'{daire.kat}. Kat'
    alt_baslik = f'Daire {daire.daire_no} - {yil} Yili Aidat Odeme Raporu'

    elements = _pdf_baslik_tablosu(apartman, alt_baslik, tarih)

    bilgi_style = ParagraphStyle('TRBilgi', fontName=PDF_FONT, fontSize=9, spaceAfter=2)
    elements.append(Paragraph(f'Daire No: {daire.daire_no}', bilgi_style))
    elements.append(Paragraph(f'Kat: {kat_str}', bilgi_style))
    elements.append(Paragraph(f'Aylik Aidat: {aidat:.2f} TL', bilgi_style))
    elements.append(Spacer(1, 8 * mm))

    rows = []
    odenen_sayisi = 0
    now = datetime.now()
    son_ay = now.month if yil == now.year else 12

    for ay in range(1, 13):
        odeme = Odeme.query.filter_by(daire_id=daire_id, yil=yil, ay=ay).first()
        odendi = odeme.odendi if odeme else False
        ay_adi = AY_ISIMLERI[ay]

        if odendi:
            durum = 'Odendi'
            odeme_tarihi = odeme.odeme_tarihi.strftime('%d.%m.%Y') if odeme.odeme_tarihi else '-'
            odenen_sayisi += 1
        elif ay <= son_ay:
            durum = 'Gecikmis'
            odeme_tarihi = '-'
        else:
            durum = 'Bekleniyor'
            odeme_tarihi = '-'

        rows.append([ay_adi, f'{aidat:.2f} TL', durum, odeme_tarihi])

    basliklar = ['Ay', 'Aidat', 'Durum', 'Odeme Tarihi']
    data = [basliklar] + rows
    col_widths = [30 * mm, 35 * mm, 35 * mm, 40 * mm]
    t = Table(data, colWidths=col_widths)

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    for i, row in enumerate(rows):
        satir_idx = i + 1
        if row[2] == 'Odendi':
            style_cmds.append(('BACKGROUND', (2, satir_idx), (2, satir_idx), colors.HexColor('#d4edda')))
        elif row[2] == 'Gecikmis':
            style_cmds.append(('BACKGROUND', (2, satir_idx), (2, satir_idx), colors.HexColor('#f8d7da')))
        else:
            style_cmds.append(('BACKGROUND', (2, satir_idx), (2, satir_idx), colors.HexColor('#e2e3e5')))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    elements.append(Spacer(1, 8 * mm))
    odenmeyen = 12 - odenen_sayisi
    ozet_style = ParagraphStyle('TROzetDaire', fontName=PDF_FONT_BOLD, fontSize=10, spaceAfter=3)
    normal_style = ParagraphStyle('TRNormalDaire', fontName=PDF_FONT, fontSize=9, spaceAfter=2)

    elements.append(Paragraph('Ozet', ozet_style))
    elements.append(Paragraph(f'Yillik Toplam: {12 * aidat:.2f} TL', normal_style))
    elements.append(Paragraph(f'Odenen: {odenen_sayisi} ay - {odenen_sayisi * aidat:.2f} TL', normal_style))
    elements.append(Paragraph(f'Kalan: {odenmeyen} ay - {odenmeyen * aidat:.2f} TL', normal_style))

    gecikmis = sum(1 for r in rows if r[2] == 'Gecikmis')
    if gecikmis > 0:
        gecikme_style = ParagraphStyle('TRGecikmeDaire', fontName=PDF_FONT_BOLD, fontSize=9, textColor=colors.HexColor('#dc3545'))
        elements.append(Paragraph(f'Gecikmis: {gecikmis} ay - {gecikmis * aidat:.2f} TL', gecikme_style))

    return _kaydet_pdf(elements)
