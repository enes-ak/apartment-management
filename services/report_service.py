import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Turkce karakter destegi icin DejaVu Sans fontu
pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
PDF_FONT = 'DejaVuSans'
PDF_FONT_BOLD = 'DejaVuSans-Bold'

from models import Apartment, Payment, Expense, ExpenseCategory, CashRegister, Setting, DuesConfig, ExtraCollection, ExtraPayment

AY_ISIMLERI = {
    1: 'Ocak', 2: 'Subat', 3: 'Mart', 4: 'Nisan', 5: 'Mayis', 6: 'Haziran',
    7: 'Temmuz', 8: 'Agustos', 9: 'Eylul', 10: 'Ekim', 11: 'Kasim', 12: 'Aralik'
}

# ── Excel Styles ──────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=14, color='FFFFFF')
SUBHEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1a1d21', end_color='1a1d21', fill_type='solid')
GREEN_FILL = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
RED_FILL = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


# ── Excel Helpers ─────────────────────────────────────────────
def _excel_header(ws, building_name, report_name, date_str):
    ws.append([building_name])
    ws.append([report_name])
    ws.append([f'Olusturma Tarihi: {date_str}'])
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=1):
        for cell in row:
            cell.font = HEADER_FONT if cell.row == 1 else SUBHEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='left')


def _excel_table_header(ws, headers, row_num):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _excel_row(ws, values, row_num, fill=None):
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if fill:
            cell.fill = fill


def _save_excel(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── PDF Helpers ───────────────────────────────────────────────
def _pdf_header_block(building_name, report_name, date_str):
    title_style = ParagraphStyle('TRTitle', fontName=PDF_FONT_BOLD, fontSize=16, spaceAfter=4)
    heading_style = ParagraphStyle('TRHeading', fontName=PDF_FONT_BOLD, fontSize=12, spaceAfter=4)
    normal_style = ParagraphStyle('TRNormal', fontName=PDF_FONT, fontSize=9, textColor=colors.grey)
    elements = []
    elements.append(Paragraph(building_name, title_style))
    elements.append(Paragraph(report_name, heading_style))
    elements.append(Paragraph(f'Olusturma Tarihi: {date_str}', normal_style))
    elements.append(Spacer(1, 10 * mm))
    return elements


def _pdf_table(headers, data_rows, col_widths=None):
    header_style = ParagraphStyle('TblH', fontName=PDF_FONT_BOLD, fontSize=10,
                                  textColor=colors.white, alignment=1, leading=12)
    cell_style = ParagraphStyle('TblC', fontName=PDF_FONT, fontSize=9,
                                alignment=1, leading=11)

    wrapped_headers = [Paragraph(str(h), header_style) for h in headers]
    wrapped_rows = []
    for row in data_rows:
        wrapped_rows.append([Paragraph(str(cell), cell_style) for cell in row])

    data = [wrapped_headers] + wrapped_rows
    t = Table(data, colWidths=col_widths)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f0f0f0')))
    t.setStyle(TableStyle(style_cmds))
    return t


def _save_pdf(elements):
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    doc.build(elements)
    output.seek(0)
    return output


def _building_name():
    return Setting.get('apartman_adi', 'Apartman')


def _date_str():
    return datetime.now().strftime('%d.%m.%Y')


def _calculate_month(year, month):
    """Calculate income/expense/breakdown for a month (independent of cash register)."""
    dues = DuesConfig.current_amount()
    paid_count = Payment.query.filter_by(year=year, month=month, is_paid=True).count()
    total_apartments = Apartment.query.count()
    dues_income = paid_count * dues

    # Extra collection income
    extra_income = 0
    collections = ExtraCollection.query.filter_by(year=year, month=month).all()
    for c in collections:
        paid = ExtraPayment.query.filter_by(collection_id=c.id, is_paid=True).count()
        extra_income += paid * c.per_unit_amount

    income = dues_income + extra_income

    expenses = Expense.query.filter_by(year=year, month=month).all()
    total_expense = sum(e.amount for e in expenses)

    category_totals = {}
    for e in expenses:
        name = e.category.category_name
        category_totals[name] = category_totals.get(name, 0) + e.amount

    return {
        'dues': dues,
        'paid_count': paid_count,
        'total_apartments': total_apartments,
        'income': income,
        'expense': total_expense,
        'category_totals': category_totals,
    }


# ══════════════════════════════════════════════════════════════
# 1) Monthly Summary
# ══════════════════════════════════════════════════════════════
def monthly_summary_excel(year, month):
    building_name = _building_name()
    month_name = AY_ISIMLERI.get(month, '')
    date_str = _date_str()
    h = _calculate_month(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Aylik Ozet'

    _excel_header(ws, building_name, f'{month_name} {year} - Aylik Ozet Raporu', date_str)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    row_num = 5
    info_items = [
        ('Aidat Tutari', f'{h["dues"]:.2f} TL'),
        ('Odeme Yapan', f'{h["paid_count"]} / {h["total_apartments"]}'),
        ('Toplam Gelir', f'{h["income"]:.2f} TL'),
        ('Toplam Gider', f'{h["expense"]:.2f} TL'),
        ('Net', f'{h["income"] - h["expense"]:.2f} TL'),
    ]
    for label, value in info_items:
        _excel_row(ws, [label, value], row_num)
        row_num += 1

    if h['category_totals']:
        row_num += 1
        _excel_table_header(ws, ['Gider Kalemi', 'Tutar'], row_num)
        row_num += 1
        for name, amount in h['category_totals'].items():
            _excel_row(ws, [name, f'{amount:.2f} TL'], row_num)
            row_num += 1
        _excel_row(ws, ['TOPLAM', f'{h["expense"]:.2f} TL'], row_num)
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).font = Font(bold=True)

    return _save_excel(wb)


def monthly_summary_pdf(year, month):
    building_name = _building_name()
    month_name = AY_ISIMLERI.get(month, '')
    date_str = _date_str()
    h = _calculate_month(year, month)

    elements = _pdf_header_block(building_name, f'{month_name} {year} - Aylik Ozet Raporu', date_str)

    summary_data = [
        ['Aidat Tutari', f'{h["dues"]:.2f} TL'],
        ['Odeme Yapan', f'{h["paid_count"]} / {h["total_apartments"]}'],
        ['Toplam Gelir', f'{h["income"]:.2f} TL'],
        ['Toplam Gider', f'{h["expense"]:.2f} TL'],
        ['Net', f'{h["income"] - h["expense"]:.2f} TL'],
    ]
    elements.append(_pdf_table(['Bilgi', 'Deger'], summary_data, col_widths=[60 * mm, 60 * mm]))
    elements.append(Spacer(1, 10 * mm))

    if h['category_totals']:
        expense_rows = [[name, f'{amount:.2f} TL'] for name, amount in h['category_totals'].items()]
        expense_rows.append(['TOPLAM', f'{h["expense"]:.2f} TL'])
        elements.append(_pdf_table(['Gider Kalemi', 'Tutar'], expense_rows, col_widths=[60 * mm, 60 * mm]))

    return _save_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 2) Payment Status
# ══════════════════════════════════════════════════════════════
def payment_status_excel(year, month):
    building_name = _building_name()
    month_name = AY_ISIMLERI.get(month, '')
    date_str = _date_str()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Odeme Durumu'

    _excel_header(ws, building_name, f'{month_name} {year} - Odeme Durumu Raporu', date_str)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    row_num = 5
    _excel_table_header(ws, ['Daire No', 'Sakin Adi', 'Durum'], row_num)
    row_num += 1

    apartments = Apartment.query.order_by(Apartment.unit_no).all()
    paid_count = 0
    for apt in apartments:
        payment = Payment.query.filter_by(apartment_id=apt.id, year=year, month=month).first()
        is_paid = payment.is_paid if payment else False
        status = 'Odendi' if is_paid else 'Odenmedi'
        fill = GREEN_FILL if is_paid else RED_FILL
        _excel_row(ws, [apt.unit_no, apt.resident_name, status], row_num, fill=fill)
        if is_paid:
            paid_count += 1
        row_num += 1

    row_num += 1
    ws.cell(row=row_num, column=1, value='Toplam Odenen:').font = Font(bold=True)
    ws.cell(row=row_num, column=2, value=f'{paid_count} / {len(apartments)}')

    return _save_excel(wb)


def payment_status_pdf(year, month):
    building_name = _building_name()
    month_name = AY_ISIMLERI.get(month, '')
    date_str = _date_str()

    elements = _pdf_header_block(building_name, f'{month_name} {year} - Odeme Durumu Raporu', date_str)

    apartments = Apartment.query.order_by(Apartment.unit_no).all()
    rows = []
    paid_count = 0
    for apt in apartments:
        payment = Payment.query.filter_by(apartment_id=apt.id, year=year, month=month).first()
        is_paid = payment.is_paid if payment else False
        status = 'Odendi' if is_paid else 'Odenmedi'
        rows.append([str(apt.unit_no), status])
        if is_paid:
            paid_count += 1

    headers = ['Daire No', 'Durum']
    data = [headers] + rows
    col_widths = [40 * mm, 50 * mm]
    t = Table(data, colWidths=col_widths)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]
    for i in range(1, len(data)):
        row_data = rows[i - 1]
        if row_data[1] == 'Odendi':
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#d4edda')))
        else:
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8d7da')))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    elements.append(Spacer(1, 5 * mm))
    summary_style = ParagraphStyle('TROzet', fontName=PDF_FONT_BOLD, fontSize=10)
    elements.append(Paragraph(f'Toplam Odenen: {paid_count} / {len(apartments)}', summary_style))

    return _save_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 3) Expense Detail
# ══════════════════════════════════════════════════════════════
def _get_expense_rows(year, month):
    """Return individual expense rows with category, description, amount."""
    expenses = Expense.query.filter_by(year=year, month=month).all()
    rows = []
    for e in expenses:
        rows.append({
            'category': e.category.category_name,
            'description': e.description or '-',
            'amount': e.amount,
        })
    total = sum(r['amount'] for r in rows)
    return rows, total


def _get_cash_balance(year, month):
    """Return the cash register balance for the given month, or None."""
    cr = CashRegister.query.filter_by(year=year, month=month).first()
    return cr


def expense_detail_excel(year, month):
    building_name = _building_name()
    month_name = AY_ISIMLERI.get(month, '')
    date_str = _date_str()
    expense_rows, total_expense = _get_expense_rows(year, month)
    cash = _get_cash_balance(year, month)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Gider Detay'

    _excel_header(ws, building_name, f'{month_name} {year} - Gider Detay Raporu', date_str)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 16

    row_num = 5
    _excel_table_header(ws, ['Gider Kalemi', 'Aciklama', 'Tutar'], row_num)
    row_num += 1
    for r in expense_rows:
        _excel_row(ws, [r['category'], r['description'], f'{r["amount"]:.2f} TL'], row_num)
        row_num += 1
    _excel_row(ws, ['TOPLAM', '', f'{total_expense:.2f} TL'], row_num)
    for col in range(1, 4):
        ws.cell(row=row_num, column=col).font = Font(bold=True)

    row_num += 2
    _excel_table_header(ws, ['Kasa Durumu', '', 'Tutar'], row_num)
    row_num += 1
    if cash:
        for label, val in [('Devir', cash.carryover), ('Gelir', cash.total_income),
                           ('Gider', cash.total_expense), ('Bakiye', cash.balance)]:
            _excel_row(ws, [label, '', f'{val:.2f} TL'], row_num)
            row_num += 1
        ws.cell(row=row_num - 1, column=1).font = Font(bold=True)
        ws.cell(row=row_num - 1, column=3).font = Font(bold=True)
    else:
        _excel_row(ws, ['Bu ay icin kasa kaydi yok', '', '-'], row_num)

    return _save_excel(wb)


def expense_detail_pdf(year, month):
    building_name = _building_name()
    month_name = AY_ISIMLERI.get(month, '')
    date_str = _date_str()
    expense_rows, total_expense = _get_expense_rows(year, month)
    cash = _get_cash_balance(year, month)

    elements = _pdf_header_block(building_name, f'{month_name} {year} - Gider Detay Raporu', date_str)

    rows = []
    for r in expense_rows:
        rows.append([r['category'], r['description'], f'{r["amount"]:.2f} TL'])
    rows.append(['TOPLAM', '', f'{total_expense:.2f} TL'])

    elements.append(_pdf_table(['Gider Kalemi', 'Aciklama', 'Tutar'], rows,
                               col_widths=[45 * mm, 70 * mm, 35 * mm]))

    elements.append(Spacer(1, 8 * mm))
    section_style = ParagraphStyle('SectionKasa', fontName=PDF_FONT_BOLD, fontSize=11, spaceAfter=4)
    elements.append(Paragraph('Kasa Durumu', section_style))
    elements.append(Spacer(1, 3 * mm))

    if cash:
        cash_rows = [
            ['Devir', f'{cash.carryover:.2f} TL'],
            ['Gelir', f'{cash.total_income:.2f} TL'],
            ['Gider', f'{cash.total_expense:.2f} TL'],
            ['Bakiye', f'{cash.balance:.2f} TL'],
        ]
        elements.append(_pdf_table(['', 'Tutar'], cash_rows,
                                   col_widths=[45 * mm, 45 * mm]))
    else:
        info_style = ParagraphStyle('KasaInfo', fontName=PDF_FONT, fontSize=9, textColor=colors.grey)
        elements.append(Paragraph('Bu ay icin kasa kaydi bulunmamaktadir.', info_style))

    return _save_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 4) Annual Summary
# ══════════════════════════════════════════════════════════════
def annual_summary_excel(year):
    building_name = _building_name()
    date_str = _date_str()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Yillik Ozet'

    _excel_header(ws, building_name, f'{year} - Yillik Ozet Raporu', date_str)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    row_num = 5
    _excel_table_header(ws, ['Ay', 'Gelir', 'Gider', 'Net'], row_num)
    row_num += 1

    total_income = 0
    total_expense = 0
    for month_num in range(1, 13):
        h = _calculate_month(year, month_num)
        total_income += h['income']
        total_expense += h['expense']
        net = h['income'] - h['expense']
        month_name = AY_ISIMLERI[month_num]
        _excel_row(ws, [month_name, f'{h["income"]:.2f} TL', f'{h["expense"]:.2f} TL',
                          f'{net:.2f} TL'], row_num)
        row_num += 1
    net_total = total_income - total_expense
    _excel_row(ws, ['TOPLAM', f'{total_income:.2f} TL', f'{total_expense:.2f} TL',
                      f'{net_total:.2f} TL'], row_num)
    for col in range(1, 5):
        ws.cell(row=row_num, column=col).font = Font(bold=True)

    return _save_excel(wb)


def annual_summary_pdf(year):
    building_name = _building_name()
    date_str = _date_str()

    elements = _pdf_header_block(building_name, f'{year} - Yillik Ozet Raporu', date_str)

    rows = []
    total_income = 0
    total_expense = 0
    for month_num in range(1, 13):
        h = _calculate_month(year, month_num)
        total_income += h['income']
        total_expense += h['expense']
        net = h['income'] - h['expense']
        month_name = AY_ISIMLERI[month_num]
        rows.append([month_name, f'{h["income"]:.2f} TL', f'{h["expense"]:.2f} TL',
                     f'{net:.2f} TL'])
    net_total = total_income - total_expense
    rows.append(['TOPLAM', f'{total_income:.2f} TL', f'{total_expense:.2f} TL',
                 f'{net_total:.2f} TL'])

    elements.append(_pdf_table(['Ay', 'Gelir', 'Gider', 'Net'], rows,
                               col_widths=[35 * mm, 40 * mm, 40 * mm, 40 * mm]))

    return _save_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 5) Apartment Payment Report
# ══════════════════════════════════════════════════════════════
def apartment_report_pdf(apartment_id, year):
    building_name = _building_name()
    date_str = _date_str()
    apartment = Apartment.query.get(apartment_id)
    dues = DuesConfig.current_amount()

    floor_str = 'Giris Kat' if apartment.floor == 0 else f'{apartment.floor}. Kat'
    subtitle = f'Daire {apartment.unit_no} - {year} Yili Aidat Odeme Raporu'

    elements = _pdf_header_block(building_name, subtitle, date_str)

    info_style = ParagraphStyle('TRBilgi', fontName=PDF_FONT, fontSize=9, spaceAfter=2)
    elements.append(Paragraph(f'Daire No: {apartment.unit_no}', info_style))
    elements.append(Paragraph(f'Kat: {floor_str}', info_style))
    elements.append(Paragraph(f'Aylik Aidat: {dues:.2f} TL', info_style))
    elements.append(Spacer(1, 8 * mm))

    rows = []
    paid_count = 0
    now = datetime.now()
    last_month = now.month if year == now.year else 12

    for month in range(1, 13):
        payment = Payment.query.filter_by(apartment_id=apartment_id, year=year, month=month).first()
        is_paid = payment.is_paid if payment else False
        month_name = AY_ISIMLERI[month]

        if is_paid:
            status = 'Odendi'
            payment_date = payment.payment_date.strftime('%d.%m.%Y') if payment.payment_date else '-'
            paid_count += 1
        elif month <= last_month:
            status = 'Gecikmis'
            payment_date = '-'
        else:
            status = 'Bekleniyor'
            payment_date = '-'

        rows.append([month_name, f'{dues:.2f} TL', status, payment_date])

    headers = ['Ay', 'Aidat', 'Durum', 'Odeme Tarihi']
    data = [headers] + rows
    col_widths = [30 * mm, 35 * mm, 35 * mm, 40 * mm]
    t = Table(data, colWidths=col_widths)

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    for i, row in enumerate(rows):
        row_idx = i + 1
        if row[2] == 'Odendi':
            style_cmds.append(('BACKGROUND', (2, row_idx), (2, row_idx), colors.HexColor('#d4edda')))
        elif row[2] == 'Gecikmis':
            style_cmds.append(('BACKGROUND', (2, row_idx), (2, row_idx), colors.HexColor('#f8d7da')))
        else:
            style_cmds.append(('BACKGROUND', (2, row_idx), (2, row_idx), colors.HexColor('#e2e3e5')))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    elements.append(Spacer(1, 8 * mm))
    unpaid = 12 - paid_count
    summary_style = ParagraphStyle('TROzetDaire', fontName=PDF_FONT_BOLD, fontSize=10, spaceAfter=3)
    normal_style = ParagraphStyle('TRNormalDaire', fontName=PDF_FONT, fontSize=9, spaceAfter=2)

    elements.append(Paragraph('Ozet', summary_style))
    elements.append(Paragraph(f'Yillik Toplam: {12 * dues:.2f} TL', normal_style))
    elements.append(Paragraph(f'Odenen: {paid_count} ay - {paid_count * dues:.2f} TL', normal_style))
    elements.append(Paragraph(f'Kalan: {unpaid} ay - {unpaid * dues:.2f} TL', normal_style))

    overdue = sum(1 for r in rows if r[2] == 'Gecikmis')
    if overdue > 0:
        overdue_style = ParagraphStyle('TRGecikmeDaire', fontName=PDF_FONT_BOLD, fontSize=9, textColor=colors.HexColor('#dc3545'))
        elements.append(Paragraph(f'Gecikmis: {overdue} ay - {overdue * dues:.2f} TL', overdue_style))

    return _save_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 6) Annual Payment Matrix (all apartments x 12 months)
# ══════════════════════════════════════════════════════════════
def payment_matrix_pdf(year):
    from reportlab.lib.pagesizes import landscape, A4 as A4_SIZE

    building_name = _building_name()
    date_str = _date_str()

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4_SIZE),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)

    title_style = ParagraphStyle('MTitle', fontName=PDF_FONT_BOLD, fontSize=14, spaceAfter=4)
    heading_style = ParagraphStyle('MHeading', fontName=PDF_FONT_BOLD, fontSize=11, spaceAfter=4)
    normal_style = ParagraphStyle('MNormal', fontName=PDF_FONT, fontSize=8, textColor=colors.grey)

    elements = []
    elements.append(Paragraph(building_name, title_style))
    elements.append(Paragraph(f'{year} - Yillik Aidat Odeme Tablosu', heading_style))
    elements.append(Paragraph(f'Olusturma Tarihi: {date_str}', normal_style))
    elements.append(Spacer(1, 6 * mm))

    apartments = Apartment.query.order_by(Apartment.unit_no).all()
    month_abbr = ['Oca', 'Sub', 'Mar', 'Nis', 'May', 'Haz',
                  'Tem', 'Agu', 'Eyl', 'Eki', 'Kas', 'Ara']
    headers = ['Daire'] + month_abbr

    rows = []
    for apt in apartments:
        row = [str(apt.unit_no)]
        for m in range(1, 13):
            payment = Payment.query.filter_by(apartment_id=apt.id, year=year, month=m).first()
            is_paid = payment.is_paid if payment else False
            row.append('\u2713' if is_paid else '\u2717')
        rows.append(row)

    data = [headers] + rows
    daire_w = 22 * mm
    month_w = (267 * mm - daire_w) / 12
    col_widths = [daire_w] + [month_w] * 12

    t = Table(data, colWidths=col_widths)

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]

    for row_idx, row in enumerate(rows):
        for col_idx in range(1, 13):
            cell_row = row_idx + 1
            if row[col_idx] == '\u2713':
                style_cmds.append(('TEXTCOLOR', (col_idx, cell_row), (col_idx, cell_row),
                                   colors.HexColor('#28a745')))
            else:
                style_cmds.append(('TEXTCOLOR', (col_idx, cell_row), (col_idx, cell_row),
                                   colors.HexColor('#dc3545')))

    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    elements.append(Spacer(1, 5 * mm))
    legend_style = ParagraphStyle('MLegend', fontName=PDF_FONT, fontSize=8, textColor=colors.HexColor('#555555'))
    elements.append(Paragraph('\u2713 = Odendi   \u2717 = Odenmedi', legend_style))

    doc.build(elements)
    output.seek(0)
    return output


# ══════════════════════════════════════════════════════════════
# 7) Cumulative Expense & Cash Status
# ══════════════════════════════════════════════════════════════
def cumulative_report_pdf(year):
    building_name = _building_name()
    date_str = _date_str()

    elements = _pdf_header_block(building_name, f'{year} - Kumulatif Gider ve Kasa Raporu', date_str)

    # --- Section 1: Monthly cumulative expense table ---
    section_style = ParagraphStyle('CRSection', fontName=PDF_FONT_BOLD, fontSize=11, spaceAfter=4)
    elements.append(Paragraph('Aylik ve Kumulatif Giderler', section_style))
    elements.append(Spacer(1, 3 * mm))

    cum_expense = 0
    cum_income = 0
    expense_rows = []
    for m in range(1, 13):
        h = _calculate_month(year, m)
        cum_expense += h['expense']
        cum_income += h['income']
        month_name = AY_ISIMLERI[m]
        expense_rows.append([
            month_name,
            f'{h["income"]:.2f} TL',
            f'{h["expense"]:.2f} TL',
            f'{cum_income:.2f} TL',
            f'{cum_expense:.2f} TL',
            f'{cum_income - cum_expense:.2f} TL',
        ])
    expense_rows.append([
        'TOPLAM',
        f'{cum_income:.2f} TL',
        f'{cum_expense:.2f} TL',
        '',
        '',
        f'{cum_income - cum_expense:.2f} TL',
    ])

    elements.append(_pdf_table(
        ['Ay', 'Gelir', 'Gider', 'Kum. Gelir', 'Kum. Gider', 'Kum. Net'],
        expense_rows,
        col_widths=[25 * mm, 28 * mm, 28 * mm, 30 * mm, 30 * mm, 30 * mm],
    ))

    elements.append(Spacer(1, 10 * mm))

    # --- Section 2: Category breakdown for the whole year ---
    elements.append(Paragraph('Yillik Gider Dagilimi (Kategoriye Gore)', section_style))
    elements.append(Spacer(1, 3 * mm))

    all_expenses = Expense.query.filter_by(year=year).all()
    cat_totals = {}
    for e in all_expenses:
        name = e.category.category_name
        cat_totals[name] = cat_totals.get(name, 0) + e.amount
    grand_total = sum(cat_totals.values())

    cat_rows = []
    for name, amount in sorted(cat_totals.items(), key=lambda x: -x[1]):
        ratio = (amount / grand_total * 100) if grand_total > 0 else 0
        cat_rows.append([name, f'{amount:.2f} TL', f'{ratio:.1f}%'])
    cat_rows.append(['TOPLAM', f'{grand_total:.2f} TL', '100%'])

    elements.append(_pdf_table(
        ['Gider Kalemi', 'Tutar', 'Oran (%)'],
        cat_rows,
        col_widths=[55 * mm, 40 * mm, 30 * mm],
    ))

    elements.append(Spacer(1, 10 * mm))

    # --- Section 3: Cash register status per month ---
    elements.append(Paragraph('Aylik Kasa Durumu', section_style))
    elements.append(Spacer(1, 3 * mm))

    cash_rows = []
    for m in range(1, 13):
        cr = CashRegister.query.filter_by(year=year, month=m).first()
        month_name = AY_ISIMLERI[m]
        if cr:
            cash_rows.append([
                month_name,
                f'{cr.carryover:.2f} TL',
                f'{cr.total_income:.2f} TL',
                f'{cr.total_expense:.2f} TL',
                f'{cr.balance:.2f} TL',
            ])
        else:
            cash_rows.append([month_name, '-', '-', '-', '-'])

    elements.append(_pdf_table(
        ['Ay', 'Devir', 'Gelir', 'Gider', 'Bakiye'],
        cash_rows,
        col_widths=[25 * mm, 35 * mm, 35 * mm, 35 * mm, 35 * mm],
    ))

    return _save_pdf(elements)


# ══════════════════════════════════════════════════════════════
# 8) Extra Collection Report
# ══════════════════════════════════════════════════════════════
def extra_collection_report_pdf(collection):
    building_name = _building_name()
    date_str = _date_str()
    month_name = AY_ISIMLERI.get(collection.month, '')

    elements = _pdf_header_block(
        building_name,
        f'Ekstra Gider Tahsilati - {collection.description}',
        date_str,
    )

    info_style = ParagraphStyle('ECInfo', fontName=PDF_FONT, fontSize=9, spaceAfter=2)
    elements.append(Paragraph(f'Donem: {month_name} {collection.year}', info_style))
    elements.append(Paragraph(f'Toplam Tutar: {collection.total_amount:.2f} TL', info_style))
    elements.append(Paragraph(f'Daire Basi: {collection.per_unit_amount:.2f} TL', info_style))
    elements.append(Spacer(1, 8 * mm))

    payments = ExtraPayment.query.filter_by(collection_id=collection.id)\
        .join(Apartment).order_by(Apartment.unit_no).all()

    rows = []
    paid_count = 0
    for ep in payments:
        if ep.is_paid:
            status = 'Odendi'
            pay_date = ep.payment_date.strftime('%d.%m.%Y') if ep.payment_date else '-'
            paid_count += 1
        else:
            status = 'Odenmedi'
            pay_date = '-'
        rows.append([str(ep.apartment.unit_no), ep.apartment.resident_name or '-', status, pay_date])

    headers = ['Daire No', 'Sakin', 'Durum', 'Odeme Tarihi']
    data = [headers] + rows
    col_widths = [25 * mm, 50 * mm, 30 * mm, 35 * mm]
    t = Table(data, colWidths=col_widths)

    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1d21')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]
    for i, row in enumerate(rows):
        row_idx = i + 1
        if row[2] == 'Odendi':
            style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#d4edda')))
        else:
            style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f8d7da')))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)

    elements.append(Spacer(1, 8 * mm))
    unpaid = len(payments) - paid_count
    collected = paid_count * collection.per_unit_amount
    remaining = unpaid * collection.per_unit_amount

    summary_style = ParagraphStyle('ECOzet', fontName=PDF_FONT_BOLD, fontSize=10, spaceAfter=3)
    normal_style = ParagraphStyle('ECNormal', fontName=PDF_FONT, fontSize=9, spaceAfter=2)
    elements.append(Paragraph('Ozet', summary_style))
    elements.append(Paragraph(f'Odenen: {paid_count} daire - {collected:.2f} TL', normal_style))
    elements.append(Paragraph(f'Odenmeyen: {unpaid} daire - {remaining:.2f} TL', normal_style))

    return _save_pdf(elements)
